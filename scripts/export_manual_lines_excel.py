"""Export all historical manually-drawn lines to Excel.

Produces a multi-sheet workbook:
  - 画线明细 (Lines) — one row per line with all geometry + metadata
  - 挂单明细 (Orders) — one row per conditional order placed on a line
  - 日历 (Calendar) — daily aggregation: # lines drawn, # orders, # filled, PnL
  - 结果标签 (Outcomes) — per-line trade outcomes (filled/cancelled/etc)

Usage:
  python scripts/export_manual_lines_excel.py

Output:
  data/reports/manual_lines_YYYYMMDD_HHMM.xlsx
  ALSO copies to ~/Desktop/manual_lines_YYYYMMDD_HHMM.xlsx

Every line value is computed with the **LOG interpolation** rule (matches
chart rendering in Log mode; see types.py::line_price_at).
"""
from __future__ import annotations

import json
import math
import shutil
from datetime import datetime, timezone, timedelta
from pathlib import Path

import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"


def _bj(ts: int | float | None) -> str | None:
    """Unix seconds → Beijing time string."""
    if ts is None:
        return None
    try:
        ts_ = float(ts)
    except (TypeError, ValueError):
        return None
    if ts_ > 1e12:  # accidentally ms
        ts_ = ts_ / 1000.0
    if ts_ <= 0:
        return None
    return (datetime.fromtimestamp(ts_, tz=timezone.utc) + timedelta(hours=8)).strftime("%Y-%m-%d %H:%M")


def _bj_date(ts: int | float | None) -> str | None:
    """Unix seconds → 'YYYY-MM-DD' in Beijing."""
    if ts is None:
        return None
    try:
        ts_ = float(ts)
    except (TypeError, ValueError):
        return None
    if ts_ > 1e12:
        ts_ = ts_ / 1000.0
    if ts_ <= 0:
        return None
    return (datetime.fromtimestamp(ts_, tz=timezone.utc) + timedelta(hours=8)).strftime("%Y-%m-%d")


def _log_line_at(price_start: float, price_end: float, t_start: int, t_end: int, ts: int) -> float:
    """Log-interpolation projection to match chart Log-mode rendering."""
    if t_end - t_start <= 0:
        return float(price_start)
    if price_start <= 0 or price_end <= 0:
        slope = (price_end - price_start) / (t_end - t_start)
        return price_start + slope * (ts - t_start)
    ratio = (ts - t_start) / (t_end - t_start)
    return math.exp(math.log(price_start) + ratio * (math.log(price_end) - math.log(price_start)))


def _slope_pct_per_hour(price_start: float, price_end: float, t_start: int, t_end: int) -> float:
    """% gain per hour along the line (log-based, i.e. true CAGR-per-hour)."""
    span = t_end - t_start
    if span <= 0 or price_start <= 0 or price_end <= 0:
        return 0.0
    total_log_change = math.log(price_end) - math.log(price_start)
    hours = span / 3600.0
    log_per_hour = total_log_change / hours
    return (math.exp(log_per_hour) - 1.0) * 100.0


def load_lines() -> list[dict]:
    p = DATA / "manual_trendlines.json"
    if not p.exists():
        return []
    with open(p, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        data = data.get("drawings") or data.get("lines") or []
    return [d for d in data if isinstance(d, dict)]


def load_conditionals() -> list[dict]:
    p = DATA / "conditional_orders.json"
    if not p.exists():
        return []
    with open(p, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        data = data.get("conditionals") or []
    return [c for c in data if isinstance(c, dict)]


def load_jsonl(name: str) -> list[dict]:
    p = DATA / name
    if not p.exists():
        return []
    out = []
    for line in open(p, "r", encoding="utf-8"):
        line = line.strip()
        if not line:
            continue
        try:
            out.append(json.loads(line))
        except Exception:
            continue
    return out


def build_lines_sheet(lines: list[dict], conds: list[dict]) -> pd.DataFrame:
    """One row per manual line with enriched fields."""
    # Index conditionals by manual_line_id for quick lookup
    conds_by_line: dict[str, list[dict]] = {}
    for c in conds:
        lid = c.get("manual_line_id")
        if lid:
            conds_by_line.setdefault(lid, []).append(c)

    rows = []
    for d in lines:
        lid = d.get("manual_line_id", "")
        t_start = int(d.get("t_start") or 0)
        t_end = int(d.get("t_end") or 0)
        p_start = float(d.get("price_start") or 0)
        p_end = float(d.get("price_end") or 0)

        span_sec = max(0, t_end - t_start)
        slope_pct_per_hour = _slope_pct_per_hour(p_start, p_end, t_start, t_end)

        # Price range
        pct_change_total = ((p_end / p_start) - 1.0) * 100.0 if p_start > 0 else 0.0

        # Projected price at current time (if extended)
        now_ts = int(datetime.now(tz=timezone.utc).timestamp())
        line_now = _log_line_at(p_start, p_end, t_start, t_end, now_ts) if t_end > t_start else p_start

        rel_conds = conds_by_line.get(lid, [])
        n_orders = len(rel_conds)
        n_filled = sum(1 for c in rel_conds if c.get("status") == "filled")
        n_cancel = sum(1 for c in rel_conds if c.get("status") == "cancelled")
        n_pending = sum(1 for c in rel_conds if c.get("status") in ("pending", "triggered"))

        rows.append({
            "线ID": lid[-20:] if lid else "",
            "创建日期 (北京)": _bj_date(d.get("created_at")),
            "币种": d.get("symbol", ""),
            "TF": d.get("timeframe", ""),
            "方向": d.get("side", ""),
            "来源": d.get("source", "manual"),
            "起点时间 (北京)": _bj(t_start),
            "起点价": p_start,
            "终点时间 (北京)": _bj(t_end),
            "终点价": p_end,
            "跨越时长 (h)": round(span_sec / 3600.0, 2),
            "总涨跌幅 %": round(pct_change_total, 3),
            "斜率 %/h (CAGR)": round(slope_pct_per_hour, 4),
            "延右": bool(d.get("extend_right", True)),
            "延左": bool(d.get("extend_left", False)),
            "当前投影价 (log)": round(line_now, 8),
            "线宽": d.get("line_width", 1.0),
            "标签": d.get("label", ""),
            "挂单总数": n_orders,
            "已成交": n_filled,
            "已取消": n_cancel,
            "挂单中": n_pending,
            "最后修改 (北京)": _bj(d.get("updated_at")),
        })

    df = pd.DataFrame(rows)
    # Sort: newest first
    if "创建日期 (北京)" in df.columns and not df.empty:
        df = df.sort_values("创建日期 (北京)", ascending=False, na_position="last")
    return df


def build_orders_sheet(conds: list[dict]) -> pd.DataFrame:
    """One row per conditional order placed."""
    rows = []
    for c in conds:
        o = c.get("order") or {}
        rows.append({
            "挂单ID": c.get("conditional_id", ""),
            "关联线ID": (c.get("manual_line_id") or "")[-20:],
            "创建日期 (北京)": _bj_date(c.get("created_at")),
            "币种": c.get("symbol", ""),
            "TF": c.get("timeframe", ""),
            "线方向": c.get("side", ""),
            "挂单方向": o.get("direction", ""),
            "类型": o.get("order_kind", ""),
            "状态": c.get("status", ""),
            "挂单时间 (北京)": _bj(c.get("created_at")),
            "触发时间 (北京)": _bj(c.get("triggered_at")),
            "取消时间 (北京)": _bj(c.get("cancelled_at")),
            "取消原因": c.get("cancel_reason", ""),
            "入场价": c.get("fill_price"),
            "数量": c.get("fill_qty"),
            "tolerance_pct": o.get("tolerance_pct_of_line"),
            "stop_offset_pct": o.get("stop_offset_pct_of_line"),
            "rr_target": o.get("rr_target"),
            "杠杆": o.get("leverage"),
            "仓位 USDT": o.get("notional_usd"),
            "反手": o.get("reverse_enabled", False),
            "交易所订单ID": c.get("exchange_order_id", ""),
            "模式": o.get("exchange_mode", ""),
        })
    df = pd.DataFrame(rows)
    if "挂单时间 (北京)" in df.columns and not df.empty:
        df = df.sort_values("挂单时间 (北京)", ascending=False, na_position="last")
    return df


def build_calendar_sheet(lines: list[dict], conds: list[dict]) -> pd.DataFrame:
    """Per-day aggregation: 画线数, 挂单数, 成交数."""
    agg: dict[str, dict] = {}

    def _slot(date: str):
        if date not in agg:
            agg[date] = {
                "日期 (北京)": date,
                "画线数": 0,
                "涉及币种": set(),
                "涉及TF": set(),
                "support线": 0,
                "resistance线": 0,
                "挂单数": 0,
                "成交数": 0,
                "取消数": 0,
                "进行中": 0,
                "long挂单": 0,
                "short挂单": 0,
            }
        return agg[date]

    for d in lines:
        date = _bj_date(d.get("created_at"))
        if not date:
            continue
        s = _slot(date)
        s["画线数"] += 1
        if d.get("symbol"):
            s["涉及币种"].add(d["symbol"])
        if d.get("timeframe"):
            s["涉及TF"].add(d["timeframe"])
        if d.get("side") == "support":
            s["support线"] += 1
        elif d.get("side") == "resistance":
            s["resistance线"] += 1

    for c in conds:
        date = _bj_date(c.get("created_at"))
        if not date:
            continue
        s = _slot(date)
        s["挂单数"] += 1
        status = c.get("status")
        if status == "filled":
            s["成交数"] += 1
        elif status == "cancelled":
            s["取消数"] += 1
        elif status in ("pending", "triggered"):
            s["进行中"] += 1
        direction = (c.get("order") or {}).get("direction")
        if direction == "long":
            s["long挂单"] += 1
        elif direction == "short":
            s["short挂单"] += 1

    rows = []
    for date in sorted(agg.keys(), reverse=True):
        s = agg[date]
        rows.append({
            "日期 (北京)": s["日期 (北京)"],
            "星期": datetime.strptime(date, "%Y-%m-%d").strftime("%A"),
            "画线数": s["画线数"],
            "涉及币种数": len(s["涉及币种"]),
            "涉及币种": ", ".join(sorted(s["涉及币种"])),
            "涉及TF": ", ".join(sorted(s["涉及TF"])),
            "support线": s["support线"],
            "resistance线": s["resistance线"],
            "挂单数": s["挂单数"],
            "成交数": s["成交数"],
            "取消数": s["取消数"],
            "进行中": s["进行中"],
            "long挂单": s["long挂单"],
            "short挂单": s["short挂单"],
        })
    return pd.DataFrame(rows)


def build_outcomes_sheet(outcomes: list[dict]) -> pd.DataFrame:
    """Outcome capture log — actually realised results."""
    rows = []
    for o in outcomes:
        rows.append({
            "线ID": (o.get("manual_line_id") or "")[-20:],
            "捕获日期 (北京)": _bj_date(o.get("capture_ts")),
            "币种": o.get("symbol", ""),
            "TF": o.get("timeframe", ""),
            "方向": o.get("direction", ""),
            "状态": o.get("status", ""),
            "已入场": bool(o.get("filled", False)),
            "入场时间 (北京)": _bj(o.get("entry_ts")),
            "入场价": o.get("entry_price"),
            "出场时间 (北京)": _bj(o.get("exit_ts")),
            "出场价": o.get("exit_price"),
            "出场原因": o.get("exit_reason"),
            "R倍": o.get("realized_r"),
            "盈亏 %": o.get("pnl_pct"),
            "持仓时长 (bars)": o.get("hold_bars"),
        })
    df = pd.DataFrame(rows)
    if "捕获日期 (北京)" in df.columns and not df.empty:
        df = df.sort_values("捕获日期 (北京)", ascending=False, na_position="last")
    return df


def build_ml_events_sheet(ml_events: list[dict]) -> pd.DataFrame:
    """All user-drawn-line ML feature events (one per draw action)."""
    rows = []
    for e in ml_events:
        rows.append({
            "时间 (北京)": _bj(e.get("ts")),
            "币种": e.get("symbol", ""),
            "TF": e.get("timeframe", ""),
            "方向": e.get("side", e.get("kind", "")),
            "阶段": e.get("capture_stage", ""),
            "起点价": e.get("price_start"),
            "终点价": e.get("price_end"),
            "起点时间ts": e.get("t_start"),
            "终点时间ts": e.get("t_end"),
            "斜率": e.get("slope"),
            "跨越bar数": e.get("bars_between"),
            "端点距离 %": e.get("anchor_distance_pct"),
            "事件": e.get("event", ""),
        })
    df = pd.DataFrame(rows)
    if "时间 (北京)" in df.columns and not df.empty:
        df = df.sort_values("时间 (北京)", ascending=False, na_position="last")
    return df


def main():
    print("[*] Loading data…")
    lines = load_lines()
    conds = load_conditionals()
    outcomes = load_jsonl("user_drawing_outcomes.jsonl")
    ml_events = load_jsonl("user_drawings_ml.jsonl")
    labels = load_jsonl("user_drawing_labels.jsonl")

    print(f"    画线:       {len(lines)} 条")
    print(f"    挂单:       {len(conds)} 条")
    print(f"    成交结果:   {len(outcomes)} 条")
    print(f"    ML 事件:    {len(ml_events)} 条")
    print(f"    学习标签:   {len(labels)} 条")

    print("[*] Building sheets…")
    df_lines = build_lines_sheet(lines, conds)
    df_orders = build_orders_sheet(conds)
    df_calendar = build_calendar_sheet(lines, conds)
    df_outcomes = build_outcomes_sheet(outcomes)
    df_ml = build_ml_events_sheet(ml_events)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir = DATA / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"manual_lines_{ts}.xlsx"

    print(f"[*] Writing {out_path}")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_calendar.to_excel(writer, sheet_name="日历", index=False)
        df_lines.to_excel(writer, sheet_name="画线明细", index=False)
        df_orders.to_excel(writer, sheet_name="挂单明细", index=False)
        df_outcomes.to_excel(writer, sheet_name="结果", index=False)
        df_ml.to_excel(writer, sheet_name="画线事件(ML)", index=False)

    # Auto-size columns
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    v = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, min(len(v), 50))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(10, max_len + 2)
        # Freeze header row
        ws.freeze_panes = "A2"
    wb.save(out_path)

    # Copy to Desktop
    desktop = Path.home() / "Desktop"
    if desktop.exists():
        dest = desktop / out_path.name
        shutil.copy2(out_path, dest)
        print(f"[✓] Copied to {dest}")
    else:
        print(f"[!] Desktop not found at {desktop}")

    print(f"\n[DONE] Report: {out_path}")
    print(f"       Size: {out_path.stat().st_size / 1024:.1f} KB")
    print(f"       Sheets: 日历 ({len(df_calendar)}), 画线明细 ({len(df_lines)}), "
          f"挂单明细 ({len(df_orders)}), 结果 ({len(df_outcomes)}), 画线事件 ({len(df_ml)})")


if __name__ == "__main__":
    main()
